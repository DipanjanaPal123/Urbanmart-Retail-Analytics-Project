import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

df = pd.read_csv("../output/fact_sales_full.csv", parse_dates=["order_date"])
cols = ["order_id", "order_date", "category", "region", "segment", "quantity", "net_sales", "profit"]
data = df[cols].copy()
data["month"] = data["order_date"].dt.strftime("%Y-%m")

wb = Workbook()
FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1D4ED8")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT, bold=True, size=16, color="1D4ED8")
LABEL_FONT = Font(name=FONT, bold=True, size=11)
thin = Side(style="thin", color="D1D5DB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---------------- Raw Data sheet ----------------
ws_data = wb.active
ws_data.title = "Raw Data"
ws_data.append(["order_id", "order_date", "category", "region", "segment", "quantity", "net_sales", "profit", "month"])
for cell in ws_data[1]:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

for row in data.itertuples(index=False):
    ws_data.append(list(row))

n_rows = len(data) + 1
for col, width in zip("ABCDEFGHI", [12, 13, 20, 10, 12, 10, 12, 12, 10]):
    ws_data.column_dimensions[col].width = width
ws_data.column_dimensions["B"].number_format = "yyyy-mm-dd"
for r in range(2, n_rows + 1):
    ws_data[f"B{r}"].number_format = "yyyy-mm-dd"
    ws_data[f"G{r}"].number_format = "$#,##0.00"
    ws_data[f"H{r}"].number_format = "$#,##0.00"

tbl = Table(displayName="RawSales", ref=f"A1:I{n_rows}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws_data.add_table(tbl)

# ---------------- KPI Dashboard sheet ----------------
ws_kpi = wb.create_sheet("Dashboard")
ws_kpi.sheet_view.showGridLines = False
ws_kpi["B2"] = "UrbanMart Retail Performance Dashboard"
ws_kpi["B2"].font = TITLE_FONT
ws_kpi["B3"] = "Data: 2022-01-01 to 2024-12-31 | Source: Raw Data (SUMIFS-driven, live)"
ws_kpi["B3"].font = Font(name=FONT, italic=True, size=9, color="6B7280")

kpi_labels = ["Total Revenue", "Total Profit", "Profit Margin %", "Total Orders", "Avg Order Value"]
kpi_formulas = [
    "=SUM('Raw Data'!G2:G{0})".format(n_rows),
    "=SUM('Raw Data'!H2:H{0})".format(n_rows),
    "=C6/C5",
    "=COUNTA('Raw Data'!A2:A{0})".format(n_rows),
    "=C5/C8",
]
kpi_formats = ["$#,##0", "$#,##0", "0.0%", "#,##0", "$#,##0.00"]
start_row = 5
for i, (label, formula, fmt) in enumerate(zip(kpi_labels, kpi_formulas, kpi_formats)):
    r = start_row + i
    ws_kpi[f"B{r}"] = label
    ws_kpi[f"B{r}"].font = LABEL_FONT
    ws_kpi[f"C{r}"] = formula
    ws_kpi[f"C{r}"].font = Font(name=FONT, bold=True, size=12, color="1D4ED8")
    ws_kpi[f"C{r}"].number_format = fmt
ws_kpi.column_dimensions["B"].width = 20
ws_kpi.column_dimensions["C"].width = 16

# ---------------- Category Summary (SUMIFS pivot) ----------------
ws_cat = wb.create_sheet("Category Summary")
ws_cat.append(["Category", "Revenue", "Profit", "Margin %", "Units Sold"])
for cell in ws_cat[1]:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

categories = sorted(data["category"].unique())
for i, cat in enumerate(categories, start=2):
    ws_cat[f"A{i}"] = cat
    ws_cat[f"B{i}"] = f"=SUMIF('Raw Data'!$C$2:$C${n_rows},A{i},'Raw Data'!$G$2:$G${n_rows})"
    ws_cat[f"C{i}"] = f"=SUMIF('Raw Data'!$C$2:$C${n_rows},A{i},'Raw Data'!$H$2:$H${n_rows})"
    ws_cat[f"D{i}"] = f"=C{i}/B{i}"
    ws_cat[f"E{i}"] = f"=SUMIF('Raw Data'!$C$2:$C${n_rows},A{i},'Raw Data'!$F$2:$F${n_rows})"
    ws_cat[f"B{i}"].number_format = "$#,##0"
    ws_cat[f"C{i}"].number_format = "$#,##0"
    ws_cat[f"D{i}"].number_format = "0.0%"
    ws_cat[f"E{i}"].number_format = "#,##0"
for col, width in zip("ABCDE", [24, 14, 14, 12, 12]):
    ws_cat.column_dimensions[col].width = width

chart1 = BarChart()
chart1.title = "Revenue by Category"
chart1.y_axis.title = "Revenue ($)"
data_ref = Reference(ws_cat, min_col=2, min_row=1, max_row=len(categories) + 1)
cats_ref = Reference(ws_cat, min_col=1, min_row=2, max_row=len(categories) + 1)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.width, chart1.height = 16, 9
ws_cat.add_chart(chart1, "G2")

# ---------------- Region Summary ----------------
ws_reg = wb.create_sheet("Region Summary")
ws_reg.append(["Region", "Revenue", "Profit", "Orders"])
for cell in ws_reg[1]:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
regions = sorted(data["region"].unique())
for i, reg in enumerate(regions, start=2):
    ws_reg[f"A{i}"] = reg
    ws_reg[f"B{i}"] = f"=SUMIF('Raw Data'!$D$2:$D${n_rows},A{i},'Raw Data'!$G$2:$G${n_rows})"
    ws_reg[f"C{i}"] = f"=SUMIF('Raw Data'!$D$2:$D${n_rows},A{i},'Raw Data'!$H$2:$H${n_rows})"
    ws_reg[f"D{i}"] = f"=COUNTIF('Raw Data'!$D$2:$D${n_rows},A{i})"
    ws_reg[f"B{i}"].number_format = "$#,##0"
    ws_reg[f"C{i}"].number_format = "$#,##0"
for col, width in zip("ABCD", [14, 14, 14, 12]):
    ws_reg.column_dimensions[col].width = width

pie = PieChart()
pie.title = "Revenue Share by Region"
data_ref = Reference(ws_reg, min_col=2, min_row=1, max_row=len(regions) + 1)
cats_ref = Reference(ws_reg, min_col=1, min_row=2, max_row=len(regions) + 1)
pie.add_data(data_ref, titles_from_data=True)
pie.set_categories(cats_ref)
pie.width, pie.height = 12, 9
ws_reg.add_chart(pie, "F2")

# ---------------- Monthly Trend ----------------
ws_month = wb.create_sheet("Monthly Trend")
ws_month.append(["Month", "Revenue"])
for cell in ws_month[1]:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
months = sorted(data["month"].unique())
for i, m in enumerate(months, start=2):
    ws_month[f"A{i}"] = m
    ws_month[f"B{i}"] = f"=SUMIF('Raw Data'!$I$2:$I${n_rows},A{i},'Raw Data'!$G$2:$G${n_rows})"
    ws_month[f"B{i}"].number_format = "$#,##0"
ws_month.column_dimensions["A"].width = 12
ws_month.column_dimensions["B"].width = 14

line = LineChart()
line.title = "Monthly Revenue Trend"
data_ref = Reference(ws_month, min_col=2, min_row=1, max_row=len(months) + 1)
cats_ref = Reference(ws_month, min_col=1, min_row=2, max_row=len(months) + 1)
line.add_data(data_ref, titles_from_data=True)
line.set_categories(cats_ref)
line.width, line.height = 20, 9
ws_month.add_chart(line, "D2")

wb.save("UrbanMart_Sales_Dashboard.xlsx")
print("Saved UrbanMart_Sales_Dashboard.xlsx")
